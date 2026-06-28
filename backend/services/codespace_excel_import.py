from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import models

MAX_CODESPACE_EXCEL_SIZE = 10 * 1024 * 1024
QUESTION_SHEET = "Questions_Import"
ANSWER_KEY_SHEET = "Answer_Key"
QUESTION_COLUMNS = {
    "ClassName", "UnitNo", "UnitTitle", "AssessmentTitle", "QuestionID", "QuestionType",
    "QuestionText", "Marks", "Difficulty", "Explanation", "StarterCode", "VisibleTestCases",
    "HiddenTestCases", "ExpectedOutput", "CaseSensitive", "Tags",
}
ANSWER_KEY_COLUMNS = {
    "AssessmentTitle", "UnitNo", "UnitTitle", "QuestionID", "QuestionType", "QuestionText",
    "CorrectAnswer", "AcceptedAnswers", "Marks", "CaseSensitive", "Explanation", "StarterCode",
    "VisibleTestCases", "HiddenTestCases", "ExpectedOutput", "EvaluationMode", "Difficulty", "Tags",
}


def text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    value = str(value).strip()
    return value or None


def parse_bool(value) -> bool:
    return (text(value) or "").casefold() in {"1", "true", "yes", "y"}


def parse_int(value, default=0) -> int:
    if text(value) is None:
        return default
    return int(float(value))


def normalize_header(value) -> str | None:
    raw = text(value)
    if not raw:
        return None
    by_compact = {name.casefold(): name for name in QUESTION_COLUMNS | ANSWER_KEY_COLUMNS}
    return by_compact.get(raw.casefold(), raw)


async def read_codespace_excel_upload(upload: UploadFile) -> bytes:
    file_name = Path((upload.filename or "").replace("\\", "/")).name
    if Path(file_name).suffix.lower() != ".xlsx":
        raise HTTPException(400, "Codespace imports must be uploaded as an .xlsx file")
    content = await upload.read(MAX_CODESPACE_EXCEL_SIZE + 1)
    await upload.close()
    if len(content) > MAX_CODESPACE_EXCEL_SIZE:
        raise HTTPException(413, "The Excel file exceeds the 10 MB limit")
    return content


def open_workbook(content: bytes):
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError) as error:
        raise HTTPException(400, "The uploaded file is not a valid .xlsx workbook") from error


def rows_from_sheet(workbook, sheet_name: str, required_columns: set[str], header_row: int | None = None):
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(400, f"The workbook must contain a {sheet_name} sheet")
    sheet = workbook[sheet_name]
    if header_row is None:
        header_row = 1
    raw_headers = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), [])
    headers = [normalize_header(value) for value in raw_headers]
    while headers and headers[-1] is None:
        headers.pop()
    missing = required_columns - set(headers)
    if missing:
        raise HTTPException(400, f"Missing required {sheet_name} columns: {', '.join(sorted(missing))}")
    if any(header is None for header in headers):
        raise HTTPException(400, f"{sheet_name} column names cannot be blank")
    if len(headers) != len(set(headers)):
        raise HTTPException(400, f"{sheet_name} column names must be unique")
    for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(text(value) for value in values):
            continue
        yield row_number, dict(zip(headers, values[:len(headers)]))


def task_title(question_id: str | None, question_text: str | None) -> str:
    if question_id:
        return f"{question_id} - Coding Task"
    return (question_text or "Coding Task")[:60]


def task_preview_from_row(row: dict, row_number: int) -> dict | None:
    question_type = (text(row.get("QuestionType")) or "").upper()
    if question_type != "CODING":
        return None
    question_id = text(row.get("QuestionID"))
    question_text = text(row.get("QuestionText"))
    if not question_text:
        raise HTTPException(400, f"Row {row_number}: QuestionText is required")
    try:
        marks = parse_int(row.get("Marks"), 10)
    except ValueError as error:
        raise HTTPException(400, f"Row {row_number}: Marks must be a number") from error
    return {
        "question_id": question_id,
        "title": task_title(question_id, question_text),
        "question_text": question_text,
        "description": question_text,
        "starter_code": text(row.get("StarterCode")),
        "expected_output": text(row.get("ExpectedOutput")),
        "marks": marks,
        "difficulty": text(row.get("Difficulty")),
        "tags": text(row.get("Tags")),
        "visible_test_cases": text(row.get("VisibleTestCases")),
        "hidden_test_cases": text(row.get("HiddenTestCases")),
        "unit_no": parse_int(row.get("UnitNo"), 0) or None,
        "unit_title": text(row.get("UnitTitle")),
        "assessment_title": text(row.get("AssessmentTitle")),
        "explanation": text(row.get("Explanation")),
        "language": "python",
    }


def answer_key_preview_from_row(row: dict, row_number: int) -> dict | None:
    question_type = (text(row.get("QuestionType")) or "").upper()
    if question_type != "CODING":
        return None
    question_id = text(row.get("QuestionID"))
    if not question_id:
        raise HTTPException(400, f"Row {row_number}: QuestionID is required")
    return {
        "question_id": question_id,
        "correct_answer": text(row.get("CorrectAnswer")),
        "accepted_answers": text(row.get("AcceptedAnswers")),
        "expected_output": text(row.get("ExpectedOutput")),
        "evaluation_mode": (text(row.get("EvaluationMode")) or "MANUAL").upper(),
        "case_sensitive": parse_bool(row.get("CaseSensitive")),
        "visible_test_cases": text(row.get("VisibleTestCases")),
        "hidden_test_cases": text(row.get("HiddenTestCases")),
        "explanation": text(row.get("Explanation")),
    }


def preview_coding_tasks(content: bytes) -> list[dict]:
    workbook = open_workbook(content)
    try:
        rows = []
        for row_number, row in rows_from_sheet(workbook, QUESTION_SHEET, QUESTION_COLUMNS):
            parsed = task_preview_from_row(row, row_number)
            if parsed:
                rows.append(parsed)
        return rows
    finally:
        workbook.close()


def preview_coding_answer_keys(content: bytes) -> list[dict]:
    workbook = open_workbook(content)
    try:
        rows = []
        for row_number, row in rows_from_sheet(workbook, ANSWER_KEY_SHEET, ANSWER_KEY_COLUMNS, header_row=4):
            parsed = answer_key_preview_from_row(row, row_number)
            if parsed:
                rows.append(parsed)
        return rows
    finally:
        workbook.close()


def import_coding_tasks(db, codespace_id: int, content: bytes) -> dict:
    workbook = open_workbook(content)
    imported = updated = skipped = 0
    errors = []
    try:
        for row_number, row in rows_from_sheet(workbook, QUESTION_SHEET, QUESTION_COLUMNS):
            question_type = (text(row.get("QuestionType")) or "").upper()
            if question_type != "CODING":
                skipped += 1
                continue
            question_id = text(row.get("QuestionID"))
            question_text = text(row.get("QuestionText"))
            if not question_text:
                skipped += 1
                errors.append(f"Row {row_number}: QuestionText is required")
                continue
            try:
                marks = parse_int(row.get("Marks"), 10)
            except ValueError:
                skipped += 1
                errors.append(f"Row {row_number}: Marks must be a number")
                continue
            task = None
            if question_id:
                task = db.query(models.CodingTask).filter_by(codespace_id=codespace_id, question_id=question_id).first()
            if not task:
                task = models.CodingTask(codespace_id=codespace_id)
                db.add(task)
                imported += 1
            else:
                updated += 1
            task.question_id = question_id
            task.unit_no = parse_int(row.get("UnitNo"), 0) or None
            task.unit_title = text(row.get("UnitTitle"))
            task.assessment_title = text(row.get("AssessmentTitle"))
            task.title = task_title(question_id, question_text)
            task.description = question_text
            task.starter_code = text(row.get("StarterCode"))
            task.expected_output = text(row.get("ExpectedOutput"))
            task.marks = marks
            task.difficulty = text(row.get("Difficulty"))
            task.explanation = text(row.get("Explanation"))
            task.visible_test_cases = text(row.get("VisibleTestCases"))
            task.hidden_test_cases = text(row.get("HiddenTestCases"))
            task.tags = text(row.get("Tags"))
            task.language = "python"
    finally:
        workbook.close()
    db.commit()
    return {"imported_count": imported, "updated_count": updated, "skipped_count": skipped, "errors": errors}


def import_coding_answer_keys(db, codespace_id: int, content: bytes) -> dict:
    workbook = open_workbook(content)
    imported = updated = skipped = 0
    errors = []
    try:
        for row_number, row in rows_from_sheet(workbook, ANSWER_KEY_SHEET, ANSWER_KEY_COLUMNS, header_row=4):
            question_type = (text(row.get("QuestionType")) or "").upper()
            if question_type != "CODING":
                skipped += 1
                continue
            question_id = text(row.get("QuestionID"))
            if not question_id:
                skipped += 1
                errors.append(f"Row {row_number}: QuestionID is required")
                continue
            task = db.query(models.CodingTask).filter_by(codespace_id=codespace_id, question_id=question_id).first()
            if not task:
                skipped += 1
                errors.append(f"Task with QuestionID {question_id} not found. Import tasks first.")
                continue
            key = task.answer_key
            if key is None:
                key = models.CodingTaskAnswerKey(task_id=task.id, question_id=question_id)
                db.add(key)
                imported += 1
            else:
                updated += 1
            key.question_id = question_id
            key.correct_answer = text(row.get("CorrectAnswer"))
            key.accepted_answers = text(row.get("AcceptedAnswers"))
            key.expected_output = text(row.get("ExpectedOutput"))
            key.evaluation_mode = (text(row.get("EvaluationMode")) or "MANUAL").upper()
            key.case_sensitive = parse_bool(row.get("CaseSensitive"))
            key.visible_test_cases = text(row.get("VisibleTestCases"))
            key.hidden_test_cases = text(row.get("HiddenTestCases"))
            key.explanation = text(row.get("Explanation"))
            task.expected_output = key.expected_output or task.expected_output
            task.visible_test_cases = key.visible_test_cases or task.visible_test_cases
            task.hidden_test_cases = key.hidden_test_cases or task.hidden_test_cases
            task.explanation = key.explanation or task.explanation
    finally:
        workbook.close()
    db.commit()
    return {"imported_count": imported, "updated_count": updated, "skipped_count": skipped, "errors": errors}

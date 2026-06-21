from io import BytesIO
import logging
from pathlib import Path
import re
from zipfile import BadZipFile

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import models


logger = logging.getLogger(__name__)
MAX_ANSWER_KEY_SIZE = 10 * 1024 * 1024
ANSWER_KEY_SHEET = "Answer_Key"
HEADER_SCAN_LIMIT = 25
COLUMN_ALIASES = {
    "questionid": "QuestionID",
    "qid": "QuestionID",
    "questiontype": "QuestionType",
    "type": "QuestionType",
    "correctanswer": "CorrectAnswer",
    "correctacceptedanswer": "CorrectAnswer",
    "answer": "CorrectAnswer",
    "acceptedanswers": "AcceptedAnswers",
    "marks": "Marks",
    "maxmarks": "Marks",
    "mark": "Marks",
    "casesensitive": "CaseSensitive",
    "explanation": "Explanation",
    "feedback": "Explanation",
    "answerexplanation": "Explanation",
    "hiddentestcases": "HiddenTestCases",
    "expectedoutput": "ExpectedOutput",
    "difficulty": "Difficulty",
    "tags": "Tags",
}
REQUIRED_COLUMNS = {"QuestionID", "QuestionType", "CorrectAnswer", "Marks"}


def value_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    value = str(value).strip()
    return value or None


def normalize_column_name(value) -> str | None:
    name = value_text(value)
    if not name:
        return None
    compact = re.sub(r"[^a-z0-9]", "", name.casefold())
    return COLUMN_ALIASES.get(compact, name)


def parse_boolean(value, row_number: int) -> tuple[bool | None, str | None]:
    raw = value_text(value)
    if raw is None:
        return None, None
    normalized = raw.casefold()
    if normalized in {"true", "yes", "1"}:
        return True, None
    if normalized in {"false", "no", "0"}:
        return False, None
    return None, f"Row {row_number}: CaseSensitive must be true, false, yes, no, 1, or 0"


def parse_marks(value, row_number: int) -> tuple[float | None, str | None]:
    if value_text(value) is None:
        return None, None
    try:
        marks = float(value)
    except (TypeError, ValueError):
        return None, f"Row {row_number}: Marks must be a number greater than or equal to 0"
    if marks < 0:
        return None, f"Row {row_number}: Marks must be a number greater than or equal to 0"
    return marks, None


def parse_answer_key_workbook(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError) as error:
        raise HTTPException(400, "The uploaded file is not a valid .xlsx workbook") from error

    try:
        if ANSWER_KEY_SHEET not in workbook.sheetnames:
            raise HTTPException(400, "The workbook must contain an Answer_Key sheet")
        sheet = workbook[ANSWER_KEY_SHEET]
        header_row = None
        headers = []
        best_headers = []
        for row_number, values in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row or HEADER_SCAN_LIMIT, HEADER_SCAN_LIMIT),
                values_only=True,
            ),
            start=1,
        ):
            candidate = [normalize_column_name(value) for value in values]
            while candidate and candidate[-1] is None:
                candidate.pop()
            if len(REQUIRED_COLUMNS.intersection(candidate)) > len(REQUIRED_COLUMNS.intersection(best_headers)):
                best_headers = candidate
            if REQUIRED_COLUMNS.issubset(candidate):
                header_row = row_number
                headers = candidate
                break
        if header_row is None:
            missing = REQUIRED_COLUMNS - set(best_headers)
            raise HTTPException(400, f"Missing required Answer_Key columns: {', '.join(sorted(missing))}")
        if any(header is None for header in headers):
            raise HTTPException(400, "Answer_Key column names cannot be blank")
        if len(headers) != len(set(headers)):
            raise HTTPException(400, "Answer_Key column names must be unique")
        logger.info("Answer key import detected header row %s with columns: %s", header_row, headers)

        parsed = []
        question_ids = set()
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            row = dict(zip(headers, values[:len(headers)]))
            if not any(value_text(value) for value in values):
                continue
            question_id = value_text(row.get("QuestionID"))
            question_type = (value_text(row.get("QuestionType")) or "").upper()
            row_errors = []
            if not question_id:
                row_errors.append(f"Row {row_number}: QuestionID is required")
            elif question_id in question_ids:
                row_errors.append(f"Row {row_number}: duplicate QuestionID {question_id}")
            else:
                question_ids.add(question_id)
            if question_type not in {"MCQ", "FILLUP", "CODING"}:
                row_errors.append(f"Row {row_number}: QuestionType must be MCQ, FILLUP, or CODING")
            marks, marks_error = parse_marks(row.get("Marks"), row_number)
            case_sensitive, case_error = parse_boolean(row.get("CaseSensitive"), row_number)
            if marks_error:
                row_errors.append(marks_error)
            if case_error:
                row_errors.append(case_error)
            parsed.append({
                "row_number": row_number,
                "order_number": len(parsed) + 1,
                "question_id": question_id,
                "question_type": question_type,
                "correct_answer": value_text(row.get("CorrectAnswer")),
                "accepted_answers": value_text(row.get("AcceptedAnswers")),
                "explanation": value_text(row.get("Explanation")),
                "hidden_test_cases": value_text(row.get("HiddenTestCases")),
                "marks": marks,
                "case_sensitive": case_sensitive,
                "difficulty": value_text(row.get("Difficulty")),
                "tags": value_text(row.get("Tags")),
                "expected_output": value_text(row.get("ExpectedOutput")),
                "errors": row_errors,
            })
    finally:
        workbook.close()

    if not parsed:
        raise HTTPException(400, "The Answer_Key sheet contains no answer-key rows")
    logger.info("Answer key import parsed %s data rows", len(parsed))
    return parsed


def validate_answer_key_rows(rows: list[dict]) -> list[str]:
    errors = []
    for row in rows:
        prefix = f"Row {row['row_number']} ({row['question_id'] or 'missing QuestionID'})"
        if row["question_type"] == "MCQ":
            answer = (row["correct_answer"] or "").strip().upper()
            if answer not in {"A", "B", "C", "D"}:
                errors.append(f"{prefix}: MCQ CorrectAnswer must be A, B, C, or D")
            else:
                row["correct_answer"] = answer
        elif row["question_type"] == "FILLUP" and not row["correct_answer"]:
            errors.append(f"{prefix}: FILLUP CorrectAnswer is required")
    return errors


def import_answer_key_for_assessment(db, assessment, rows: list[dict]) -> dict:
    validation_errors = set(validate_answer_key_rows(rows))
    by_excel_id = {question.question_id_from_excel: question for question in assessment.questions}
    by_order = {question.order_number: question for question in assessment.questions}
    matched_question_ids = set()
    imported = 0
    updated = 0
    marks_updated = 0
    skipped = 0
    missing_questions = []
    errors = []

    for row in rows:
        prefix = f"Row {row['row_number']} ({row['question_id']})"
        row_errors = list(row.get("errors", []))
        row_errors.extend(error for error in validation_errors if error.startswith(prefix))
        if row_errors:
            skipped += 1
            errors.extend(row_errors)
            continue

        question = by_excel_id.get(row["question_id"]) or by_order.get(row["order_number"])
        if not question:
            skipped += 1
            missing_id = row["question_id"] or f"row {row['order_number']}"
            missing_questions.append(missing_id)
            reason = f"Question not found for QuestionID: {missing_id}"
            errors.append(reason)
            logger.debug("Answer key import skipped %s: %s", prefix, reason)
            continue
        if question.id in matched_question_ids:
            skipped += 1
            errors.append(f"{prefix}: matches a question already updated by another row")
            continue
        matched_question_ids.add(question.id)
        question.question_type = row["question_type"]
        if row["marks"] is not None:
            question.marks = row["marks"]
            marks_updated += 1
        if row["case_sensitive"] is not None:
            question.case_sensitive = row["case_sensitive"]
        if row["difficulty"] is not None:
            question.difficulty = row["difficulty"]
        if row["tags"] is not None:
            question.tags = row["tags"]
        if row["expected_output"] is not None:
            question.expected_output = row["expected_output"]
        key = question.answer_key
        if key is None:
            key = models.AssessmentAnswerKey(question_id=question.id)
            question.answer_key = key
            db.add(key)
            imported += 1
        else:
            updated += 1
        key.correct_answer = row["correct_answer"]
        key.accepted_answers = row["accepted_answers"]
        key.explanation = row["explanation"]
        key.hidden_test_cases = row["hidden_test_cases"]

    total_marks = sum(question.marks for question in assessment.questions)
    for attempt in assessment.attempts:
        attempt.total_marks = total_marks
    db.commit()
    total_questions = len(assessment.questions)
    total_answer_keys = sum(
        (
            question.question_type == "MCQ"
            and question.answer_key is not None
            and (question.answer_key.correct_answer or "").strip().upper() in {"A", "B", "C", "D"}
        ) or (
            question.question_type == "FILLUP"
            and question.answer_key is not None
            and bool((question.answer_key.correct_answer or "").strip())
        ) or (
            question.question_type == "CODING"
            and question.answer_key is not None
        )
        for question in assessment.questions
    )
    summary = {
        "imported": imported,
        "updated": updated,
        "marks_updated": marks_updated,
        "skipped": skipped,
        "missing_questions": missing_questions,
        "errors": errors,
        "total_answer_keys": total_answer_keys,
        "total_questions": total_questions,
        "missing_answer_keys": total_questions - total_answer_keys,
    }
    logger.info(
        "Answer key import matched=%s inserted=%s updated=%s marks_updated=%s skipped=%s",
        len(matched_question_ids), imported, updated, marks_updated, skipped,
    )
    return summary


async def read_answer_key_upload(upload: UploadFile) -> bytes:
    file_name = Path((upload.filename or "").replace("\\", "/")).name
    if Path(file_name).suffix.lower() != ".xlsx":
        raise HTTPException(400, "Answer keys must be uploaded as an .xlsx file")
    content = await upload.read(MAX_ANSWER_KEY_SIZE + 1)
    await upload.close()
    if len(content) > MAX_ANSWER_KEY_SIZE:
        raise HTTPException(413, "The answer-key workbook exceeds the 10 MB limit")
    return content

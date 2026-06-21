from io import BytesIO
from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile

import aiofiles
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

BACKEND_ROOT = Path(__file__).resolve().parent
ASSESSMENT_UPLOAD_ROOT = BACKEND_ROOT / "uploads" / "assessments"
MAX_EXCEL_SIZE = 10 * 1024 * 1024
REQUIRED_COLUMNS = {"AssessmentTitle", "QuestionID", "QuestionType", "QuestionText", "Marks"}
SUPPORTED_COLUMNS = {
    "ClassName", "UnitNo", "UnitTitle", "AssessmentTitle", "QuestionID", "QuestionType",
    "QuestionText", "OptionA", "OptionB", "OptionC", "OptionD", "CorrectAnswer",
    "AcceptedAnswers", "Marks", "Difficulty", "Explanation", "StarterCode",
    "VisibleTestCases", "HiddenTestCases", "ExpectedOutput", "CaseSensitive", "Tags",
}
COLUMN_NAMES = {name.casefold(): name for name in SUPPORTED_COLUMNS}
PREFERRED_QUESTION_SHEET = "Questions_Import"
HEADER_SCAN_LIMIT = 25


def text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    value = str(value).strip()
    return value or None


def parse_bool(value) -> bool:
    return (text(value) or "").lower() in {"1", "true", "yes", "y"}


def canonical_header(value) -> str | None:
    name = text(value)
    return COLUMN_NAMES.get(name.casefold(), name) if name else None


def find_question_table(workbook):
    """Find the template's question table instead of assuming the active sheet."""
    preferred = workbook[PREFERRED_QUESTION_SHEET] if PREFERRED_QUESTION_SHEET in workbook.sheetnames else None
    sheets = ([preferred] if preferred else []) + [sheet for sheet in workbook.worksheets if sheet is not preferred]
    best_headers = []
    for sheet in sheets:
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or HEADER_SCAN_LIMIT, HEADER_SCAN_LIMIT), values_only=True),
            start=1,
        ):
            headers = [canonical_header(value) for value in values]
            while headers and headers[-1] is None:
                headers.pop()
            if len(REQUIRED_COLUMNS.intersection(headers)) > len(REQUIRED_COLUMNS.intersection(best_headers)):
                best_headers = headers
            if REQUIRED_COLUMNS.issubset(headers):
                return sheet, row_number, headers
    missing = REQUIRED_COLUMNS - set(best_headers)
    raise HTTPException(400, f"Missing required Excel columns: {', '.join(sorted(missing))}")


def parse_workbook(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError) as error:
        raise HTTPException(400, "The uploaded file is not a valid .xlsx workbook") from error

    sheet, header_row, header_names = find_question_table(workbook)
    if not header_names or any(header is None for header in header_names):
        workbook.close()
        raise HTTPException(400, "Excel column names cannot be blank")
    if len(header_names) != len(set(header_names)):
        workbook.close()
        raise HTTPException(400, "Excel column names must be unique")
    rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)

    parsed = []
    question_ids = set()
    workbook_titles = set()
    for row_number, values in enumerate(rows, start=header_row + 1):
        row = dict(zip(header_names, values[:len(header_names)]))
        if not any(value is not None for value in values):
            continue
        if not any(text(row.get(column)) for column in ("QuestionID", "QuestionType", "QuestionText", "Marks")):
            continue
        question_id = text(row.get("QuestionID"))
        question_type = (text(row.get("QuestionType")) or "").upper()
        question_text = text(row.get("QuestionText"))
        row_title = text(row.get("AssessmentTitle"))
        if not question_id or not question_text:
            raise HTTPException(400, f"Row {row_number}: QuestionID and QuestionText are required")
        if question_id in question_ids:
            raise HTTPException(400, f"Row {row_number}: duplicate QuestionID {question_id}")
        question_ids.add(question_id)
        if question_type not in {"MCQ", "FILLUP", "CODING"}:
            raise HTTPException(400, f"Row {row_number}: QuestionType must be MCQ, FILLUP, or CODING")
        if row_title:
            workbook_titles.add(row_title.casefold())
            if len(workbook_titles) > 1:
                raise HTTPException(400, f"Row {row_number}: AssessmentTitle must be consistent across all question rows")
        try:
            marks = float(row.get("Marks"))
        except (TypeError, ValueError) as error:
            raise HTTPException(400, f"Row {row_number}: Marks must be a positive number") from error
        if marks <= 0:
            raise HTTPException(400, f"Row {row_number}: Marks must be a positive number")

        options = {letter: text(row.get(f"Option{letter}")) for letter in "ABCD"}
        correct_answer = text(row.get("CorrectAnswer"))
        if question_type == "MCQ":
            if not all(options.values()):
                raise HTTPException(400, f"Row {row_number}: MCQ options A-D are required")
            correct_answer = (correct_answer or "").upper()
            if correct_answer not in {"A", "B", "C", "D"}:
                raise HTTPException(400, f"Row {row_number}: MCQ CorrectAnswer must be A, B, C, or D")
        elif question_type == "FILLUP" and not correct_answer:
            raise HTTPException(400, f"Row {row_number}: FILLUP CorrectAnswer is required")

        parsed.append({
            "question": {
                "question_id_from_excel": question_id,
                "question_type": question_type,
                "question_text": question_text,
                "option_a": options["A"], "option_b": options["B"],
                "option_c": options["C"], "option_d": options["D"],
                "marks": marks,
                "difficulty": text(row.get("Difficulty")),
                "starter_code": text(row.get("StarterCode")),
                "visible_test_cases": text(row.get("VisibleTestCases")),
                "expected_output": text(row.get("ExpectedOutput")),
                "case_sensitive": parse_bool(row.get("CaseSensitive")),
                "tags": text(row.get("Tags")),
                "order_number": len(parsed) + 1,
            },
            "answer_key": {
                "correct_answer": correct_answer,
                "accepted_answers": text(row.get("AcceptedAnswers")),
                "explanation": text(row.get("Explanation")),
                "hidden_test_cases": text(row.get("HiddenTestCases")),
            },
        })
    workbook.close()
    if not parsed:
        raise HTTPException(400, "The workbook contains no question rows")
    return parsed


async def read_excel_upload(upload: UploadFile) -> bytes:
    file_name = Path((upload.filename or "").replace("\\", "/")).name
    if Path(file_name).suffix.lower() != ".xlsx":
        raise HTTPException(400, "Assessment questions must be uploaded as an .xlsx file")
    content = await upload.read(MAX_EXCEL_SIZE + 1)
    await upload.close()
    if len(content) > MAX_EXCEL_SIZE:
        raise HTTPException(413, "The Excel file exceeds the 10 MB limit")
    return content


async def store_excel_source(content: bytes, assessment_id: int) -> str:
    directory = ASSESSMENT_UPLOAD_ROOT / str(assessment_id)
    directory.mkdir(parents=True, exist_ok=True)
    file_name = f"{uuid4().hex}.xlsx"
    path = directory / file_name
    async with aiofiles.open(path, "wb") as target:
        await target.write(content)
    return path.relative_to(BACKEND_ROOT).as_posix()


def remove_excel_source(relative_path: str | None) -> None:
    if not relative_path:
        return
    path = (BACKEND_ROOT / relative_path).resolve()
    root = ASSESSMENT_UPLOAD_ROOT.resolve()
    if root in path.parents:
        path.unlink(missing_ok=True)
        try:
            path.parent.rmdir()
        except OSError:
            pass
